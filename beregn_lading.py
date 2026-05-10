#!/usr/bin/env python3
"""
Elbillading - strømkostnad beregning, Køhnkvartalet Sameie

Leser CloudCharge CSV-filer fra mappen CloudCharge/ og strømpris-PDF-er
fra mappen Faktura/, beregner kostnad per ladepunkt per måned og skriver
ut Excel-fil til regnskapsfører.

Bruk:
    python beregn_lading.py

Mappestruktur:
    CloudCharge/   - CSV-filer lastet ned fra CloudCharge-portalen
    Faktura/       - PDF-fakturaer fra strømleverandør (Ustekveikja Energi)

Måneder uten registrert strømpris utelates automatisk fra rapporten.
Strømpriser leses fra PDF-fakturaer. Eldre måneder uten PDF-faktura
hentes fra Excel-fanen Strømpriser som fallback.
"""

import argparse
import os
import glob
import re
from datetime import datetime

import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MAPPE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FIL = os.path.join(MAPPE, "Elbillading Strøm+Beboere.xlsx")
CLOUDCHARGE_MAPPE = os.path.join(MAPPE, "CloudCharge")
FAKTURA_MAPPE = os.path.join(MAPPE, "Faktura")

MÅNEDER = {
    1: "Januar", 2: "Februar", 3: "Mars", 4: "April",
    5: "Mai", 6: "Juni", 7: "Juli", 8: "August",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}

# Påslag på spotpris for å dekke usikkerhet rundt ladetidspunkt på døgnet
PRISPÅSLAG = 1.20


# ── Strømpriser ────────────────────────────────────────────────────────────

def _sf(s):
    """Norsk desimalstreng → float ('2 895,01' → 2895.01)."""
    return float(str(s).replace("\xa0", "").replace(" ", "").replace(",", "."))


FORVENTET_ANLEGGSREFERANSE = "Strøm Fellesareal"


def les_pris_fra_pdf(filsti):
    """
    Beregner totalpris per kWh som: Total fakturabeløp inkl. MVA / Strømpris kWh × 100.
    Returnerer (år, måned, totalt_pr_kwh) eller None ved parsefeil eller feil anleggsreferanse.
    """
    filnavn = os.path.basename(filsti)
    with pdfplumber.open(filsti) as pdf:
        tekst = "\n".join(p.extract_text() or "" for p in pdf.pages)

    # Sjekk at fakturaen gjelder riktig anlegg (Strøm Fellesareal)
    m_anlegg = re.search(r"Anleggsreferanse\s*[:\-]?\s*(.+)", tekst)
    if not m_anlegg:
        print(f"  ! {filnavn}: Fant ikke Anleggsreferanse — hopper over.")
        return None
    anleggsref = m_anlegg.group(1).strip()
    if FORVENTET_ANLEGGSREFERANSE.lower() not in anleggsref.lower():
        print(f"  ! {filnavn}: Anleggsreferanse er '{anleggsref}' "
              f"(forventet '{FORVENTET_ANLEGGSREFERANSE}') — hopper over.")
        return None

    # Periode og forbruk fra Strømpris-linjen
    m = re.search(
        r"Strømpris\s+(\d{2})\.(\d{2})\.(\d{2})-(\d{2})\.(\d{2})\.(\d{2})\s+"
        r"([\d ,]+,\d+)\s+kWh",
        tekst,
    )
    if not m:
        print(f"  ! {filnavn}: Fant ikke faktureringsperiode — hopper over.")
        return None
    måned_nr = int(m.group(2))
    år = 2000 + int(m.group(3))
    slutt_måned = int(m.group(5))
    slutt_år = 2000 + int(m.group(6))
    forbruk_kwh = _sf(m.group(7))

    # Normal faktura: start og slutt er én måned fra hverandre (f.eks. 01.04.26–01.05.26).
    if (slutt_år * 12 + slutt_måned) - (år * 12 + måned_nr) > 1:
        print(f"  ! {filnavn}: Perioden dekker mer enn én måned — kontroller fakturaen manuelt.")
        return None

    # Total fakturabeløp inkl. MVA — "å betale"-linjen finnes i alle fakturaformater
    m = re.search(r"[Åå] betale\s+(\d[\d \xa0]*,\d{2})\s+kr", tekst)
    if not m:
        print(f"  ! {filnavn}: Fant ikke totalbeløp — hopper over.")
        return None
    total_kr = _sf(m.group(1))

    totalt_pr_kwh = round(total_kr / forbruk_kwh * 100, 4)
    print(f"  PDF {filnavn}: {MÅNEDER[måned_nr]} {år} — "
          f"{total_kr:.2f} kr / {forbruk_kwh:.2f} kWh = {totalt_pr_kwh:.4f} øre/kWh "
          f"(anlegg: {anleggsref})")
    return (år, måned_nr, totalt_pr_kwh)


def les_strompriser(faktura_mappe):
    """
    Leser strømpriser fra PDF-fakturaer (primær) og Excel (fallback).
    PDF-priser overstyrer Excel for samme måned.
    Skriver ut hvilken kilde som brukes per måned.
    """
    # Fallback: Excel-fanen Strømpriser
    excel_priser = {}
    wb = openpyxl.load_workbook(EXCEL_FIL, data_only=True)
    ws = wb["Strømpriser"]
    for row in ws.iter_rows(values_only=True):
        dato, pris = row[0], row[10]
        if isinstance(dato, datetime) and isinstance(pris, (int, float)):
            excel_priser[(dato.year, dato.month)] = pris

    # Primær: PDF-fakturaer i Faktura-mappen
    pdf_priser = {}
    pdf_filer = sorted(glob.glob(os.path.join(faktura_mappe, "*.pdf")))
    pdf_feil = []
    for filsti in pdf_filer:
        resultat = les_pris_fra_pdf(filsti)
        if resultat:
            år, måned_nr, pris = resultat
            pdf_priser[(år, måned_nr)] = pris
        else:
            pdf_feil.append(os.path.basename(filsti))

    if pdf_feil:
        print(f"  ! Kunne ikke lese pris fra: {', '.join(pdf_feil)}")

    # Meld fra om kilde per måned
    alle_nøkler = set(excel_priser) | set(pdf_priser)
    priser = {}
    for nøkkel in sorted(alle_nøkler):
        if nøkkel in pdf_priser:
            priser[nøkkel] = pdf_priser[nøkkel]
        else:
            priser[nøkkel] = excel_priser[nøkkel]

    print(f"  Strømpriser lastet: {len(pdf_priser)} fra PDF, "
          f"{len(alle_nøkler) - len(pdf_priser)} fra Excel")

    return priser


# ── Beboere ────────────────────────────────────────────────────────────────

def les_beboere():
    wb = openpyxl.load_workbook(EXCEL_FIL, data_only=True)
    ws = wb["Beboere"]
    beboere = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        nr, navn, hnummer = row[0], row[1], row[2]
        if nr and navn:
            beboere[int(nr)] = {
                "navn": str(navn).strip(),
                "hnummer": str(hnummer).strip() if hnummer else "",
            }
    return beboere


# ── CloudCharge CSV ────────────────────────────────────────────────────────

def norsk_float(verdi):
    if pd.isna(verdi):
        return 0.0
    if isinstance(verdi, (int, float)):
        return float(verdi)
    return float(str(verdi).replace("\xa0", "").replace(" ", "").replace(",", "."))


def les_csv_filer(cloudcharge_mappe):
    filer = sorted(glob.glob(os.path.join(cloudcharge_mappe, "*.csv")))
    if not filer:
        raise FileNotFoundError(
            f"Ingen CSV-filer funnet i {cloudcharge_mappe}. "
            "Legg CloudCharge-fil(er) i mappen og prøv igjen."
        )

    print(f"  {len(filer)} CSV-fil(er) fra CloudCharge:")
    for f in filer:
        print(f"    {os.path.basename(f)}")

    deler = [pd.read_csv(f, sep=";", encoding="utf-8-sig", dtype=str) for f in filer]
    return pd.concat(deler, ignore_index=True)


# ── Summering ──────────────────────────────────────────────────────────────

def legg_til_summer(resultater):
    """Injiserer en Sum-rad etter hver gruppe av måneder per ladepunkt."""
    output = []
    i = 0
    while i < len(resultater):
        ladepunkt = resultater[i]["Ladepunkt"]
        gruppe = []
        while i < len(resultater) and resultater[i]["Ladepunkt"] == ladepunkt:
            gruppe.append(resultater[i])
            i += 1

        output.extend(gruppe)

        if len(gruppe) > 1:
            output.append({
                "Ladepunkt": ladepunkt,
                "Navn": gruppe[0]["Navn"],
                "Leilighet": gruppe[0]["Leilighet"],
                "År": "Sum",
                "Måned": None,
                "Forbruk (kWh)": round(sum(r["Forbruk (kWh)"] for r in gruppe), 2),
                "Strømpris inkl. 20% påslag (øre/kWh)": None,
                "Strømkost (kr)": round(sum(r["Strømkost (kr)"] for r in gruppe), 2),
                "_summary": True,
            })

    return output


# ── Hovedlogikk ────────────────────────────────────────────────────────────

def beregn(fra=None, til=None, faktura_mappe=None, cloudcharge_mappe=None):
    faktura_mappe = faktura_mappe or FAKTURA_MAPPE
    cloudcharge_mappe = cloudcharge_mappe or CLOUDCHARGE_MAPPE

    if fra and til:
        print(f"Periode: {fra[0]}.{fra[1]:02d} – {til[0]}.{til[1]:02d}")
    print(f"Faktura-mappe    : {faktura_mappe}")
    print(f"CloudCharge-mappe: {cloudcharge_mappe}")
    print("Leser strømpriser (PDF + Excel)...")
    priser = les_strompriser(faktura_mappe)

    print("Leser beboerregister fra Excel...")
    beboere = les_beboere()

    print("Leser CloudCharge CSV...")
    data = les_csv_filer(cloudcharge_mappe)

    data["Energi_kWh"] = data["Energi (kWh)"].apply(norsk_float)
    data["Sluttdato_dt"] = pd.to_datetime(data["Sluttdato"], format="%Y-%m-%d", errors="coerce")
    data["Uttak"] = pd.to_numeric(data["Uttak nummer"], errors="coerce")
    data["BeboerNr"] = data["Uttak"] + 100
    data = data.dropna(subset=["Sluttdato_dt", "Uttak"])

    # Sluttdato for månedstildeling — samsvarer med CloudCharges eksportformat
    # (CloudCharge inkluderer sesjoner i perioden basert på sluttdato).
    data["År"] = data["Sluttdato_dt"].dt.year.astype(int)
    data["Måned"] = data["Sluttdato_dt"].dt.month.astype(int)
    data["BeboerNr"] = data["BeboerNr"].astype(int)

    måneder_i_csv = sorted({(int(r["År"]), int(r["Måned"])) for _, r in data.iterrows()})
    if fra and til:
        måneder_i_csv = [m for m in måneder_i_csv if fra <= m <= til]
    gyldige_måneder = [m for m in måneder_i_csv if m in priser]
    utelatte_måneder = [m for m in måneder_i_csv if m not in priser]

    forbruk_data = data[data["Energi_kWh"] > 0]
    gruppert = forbruk_data.groupby(["BeboerNr", "År", "Måned"])["Energi_kWh"].sum()

    resultater = []

    for beboer_nr in sorted(beboere.keys()):
        beboer = beboere[beboer_nr]
        for (år, måned) in gyldige_måneder:
            forbruk = round(gruppert.get((beboer_nr, år, måned), 0.0), 2)
            pris_ore = round(priser[(år, måned)] * PRISPÅSLAG, 4)
            kostnad = round(forbruk * pris_ore / 100, 2)

            resultater.append({
                "Ladepunkt": beboer_nr,
                "Navn": beboer["navn"],
                "Leilighet": beboer["hnummer"],
                "År": år,
                "Måned": MÅNEDER[måned],
                "_sort": (beboer_nr, år, måned),
                "_summary": False,
                "Forbruk (kWh)": forbruk,
                "Strømpris inkl. 20% påslag (øre/kWh)": pris_ore,
                "Strømkost (kr)": kostnad,
            })

    if utelatte_måneder:
        utelatt_str = ", ".join(f"{MÅNEDER[m]} {å}" for å, m in utelatte_måneder)
        print(f"\n  Følgende måneder utelatt (ingen PDF-faktura eller Excel-pris): {utelatt_str}")

    if not resultater:
        print("\nIngen resultater å skrive. Sjekk at filer er lagt i riktige mapper.")
        return

    resultater.sort(key=lambda x: x["_sort"])
    for r in resultater:
        del r["_sort"]

    resultater = legg_til_summer(resultater)

    tidsstempel = datetime.now().strftime("%Y%m%d_%H%M")
    if fra and til:
        periode_str = f"{fra[0]}{fra[1]:02d}-{til[0]}{til[1]:02d}_"
    else:
        periode_str = ""
    output_fil = os.path.join(MAPPE, f"Fakturering_{periode_str}{tidsstempel}.xlsx")
    første_måned = gyldige_måneder[0]
    siste_måned = gyldige_måneder[-1]
    skriv_excel(resultater, output_fil, første_måned, siste_måned, beboere)

    rader = [r for r in resultater if not r["_summary"]]
    print(f"\nTotalt forbruk : {sum(r['Forbruk (kWh)'] for r in rader):,.2f} kWh")
    print(f"Total kostnad  : kr {sum(r['Strømkost (kr)'] for r in rader):,.2f}")
    print(f"\nFerdig! Fil lagret: {os.path.basename(output_fil)}")


# ── Excel-output ───────────────────────────────────────────────────────────

def skriv_excel(resultater, filsti, første_måned, siste_måned, beboere):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fakturering"

    kolonner = [
        "Ladepunkt", "Navn", "Leilighet", "År", "Måned",
        "Forbruk (kWh)", "Strømpris inkl. 20% påslag (øre/kWh)", "Strømkost (kr)",
    ]
    kolonnebredder = [12, 32, 12, 8, 14, 16, 32, 16]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    summary_font = Font(bold=True)
    summary_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    topp_kant = Border(top=Side(style="thin"))
    annenhver_fill = PatternFill(start_color="EEF3F9", end_color="EEF3F9", fill_type="solid")

    for col, navn in enumerate(kolonner, 1):
        celle = ws.cell(row=1, column=col, value=navn)
        celle.font = header_font
        celle.fill = header_fill
        celle.alignment = Alignment(horizontal="center")

    forrige_ladepunkt = None
    gruppe_teller = 0

    for rad_nr, rad in enumerate(resultater, 2):
        er_summary = rad.get("_summary", False)

        if rad["Ladepunkt"] != forrige_ladepunkt:
            forrige_ladepunkt = rad["Ladepunkt"]
            gruppe_teller += 1

        fyll = summary_fill if er_summary else (annenhver_fill if gruppe_teller % 2 == 0 else None)

        def sett(col, verdi, tallformat=None):
            c = ws.cell(row=rad_nr, column=col, value=verdi)
            if fyll:
                c.fill = fyll
            if tallformat:
                c.number_format = tallformat
            if er_summary:
                c.font = summary_font
                c.border = topp_kant
            return c

        sett(1, rad["Ladepunkt"])
        sett(2, rad["Navn"])
        sett(3, rad["Leilighet"])
        sett(4, rad["År"])
        sett(5, rad["Måned"])
        sett(6, rad["Forbruk (kWh)"], "#,##0.00")
        sett(7, rad["Strømpris inkl. 20% påslag (øre/kWh)"], "#,##0.00")
        sett(8, rad["Strømkost (kr)"], "#,##0.00")

    for col, bredde in enumerate(kolonnebredder, 1):
        ws.column_dimensions[get_column_letter(col)].width = bredde

    ws.freeze_panes = "A2"

    # ── Fane 2: Oppsummert ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Oppsummert")

    periode_tekst = (
        f"Fakturaunderlag for perioden "
        f"{MÅNEDER[første_måned[1]]} {første_måned[0]} "
        f"til {MÅNEDER[siste_måned[1]]} {siste_måned[0]}"
    )
    ws2.cell(row=1, column=1, value=periode_tekst).font = Font(bold=True, size=12)
    ws2.merge_cells("A1:E1")

    kol2 = ["Ladepunkt", "Navn", "Leilighet", "Forbruk (kWh)", "Strømkost (kr)"]
    bredder2 = [12, 32, 12, 16, 16]
    for col, navn in enumerate(kol2, 1):
        c = ws2.cell(row=3, column=col, value=navn)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    # Summer forbruk og kostnad per ladepunkt — alle ladepunkter tas med, også de med null forbruk
    summer = {}
    for rad in resultater:
        if rad.get("_summary"):
            continue
        lp = rad["Ladepunkt"]
        if lp not in summer:
            summer[lp] = {"Navn": rad["Navn"], "Leilighet": rad["Leilighet"],
                          "Forbruk (kWh)": 0.0, "Strømkost (kr)": 0.0}
        summer[lp]["Forbruk (kWh)"] += rad["Forbruk (kWh)"]
        summer[lp]["Strømkost (kr)"] += rad["Strømkost (kr)"]
    # Ladepunkter som ikke finnes i resultater (ingen gyldige måneder) hentes fra beboere
    for lp, beboer in beboere.items():
        if lp not in summer:
            summer[lp] = {"Navn": beboer["navn"], "Leilighet": beboer["hnummer"],
                          "Forbruk (kWh)": 0.0, "Strømkost (kr)": 0.0}

    rad_nr2 = 4
    for lp in sorted(summer.keys()):
        s = summer[lp]
        ws2.cell(row=rad_nr2, column=1, value=lp)
        ws2.cell(row=rad_nr2, column=2, value=s["Navn"])
        ws2.cell(row=rad_nr2, column=3, value=s["Leilighet"])
        c_kwh = ws2.cell(row=rad_nr2, column=4, value=round(s["Forbruk (kWh)"], 2))
        c_kwh.number_format = "#,##0.00"
        c_kr = ws2.cell(row=rad_nr2, column=5, value=round(s["Strømkost (kr)"], 2))
        c_kr.number_format = "#,##0.00"
        rad_nr2 += 1

    # Totallinje
    ws2.cell(row=rad_nr2, column=3, value="Totalt").font = Font(bold=True)
    c_tot_kwh = ws2.cell(row=rad_nr2, column=4,
                         value=round(sum(s["Forbruk (kWh)"] for s in summer.values()), 2))
    c_tot_kwh.number_format = "#,##0.00"
    c_tot_kwh.font = Font(bold=True)
    c_tot_kr = ws2.cell(row=rad_nr2, column=5,
                        value=round(sum(s["Strømkost (kr)"] for s in summer.values()), 2))
    c_tot_kr.number_format = "#,##0.00"
    c_tot_kr.font = Font(bold=True)
    for col in range(1, 6):
        ws2.cell(row=rad_nr2, column=col).border = Border(top=Side(style="thin"))

    for col, bredde in enumerate(bredder2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = bredde

    wb.save(filsti)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Beregn strømkostnader for elbil-ladepunkter og generer Excel-rapport.",
        epilog="Eksempel: python beregn_lading.py 2026.01 2026.04 -F /data/Faktura -C /data/CSV",
    )
    parser.add_argument(
        "fra", nargs="?", metavar="YYYY.MM",
        help="Fra-måned (inklusiv). Må oppgis sammen med TIL.",
    )
    parser.add_argument(
        "til", nargs="?", metavar="YYYY.MM",
        help="Til-måned (inklusiv). Må oppgis sammen med FRA.",
    )
    parser.add_argument(
        "-F", "--faktura", metavar="MAPPE",
        help=f"Sti til mappe med PDF-fakturaer (standard: {FAKTURA_MAPPE})",
    )
    parser.add_argument(
        "-C", "--cloudcharge", metavar="MAPPE",
        help=f"Sti til mappe med CloudCharge CSV-filer (standard: {CLOUDCHARGE_MAPPE})",
    )
    args = parser.parse_args()

    # Valider periode
    fra = til = None
    if args.fra or args.til:
        if not (args.fra and args.til):
            parser.error("Både FRA og TIL må oppgis, eller ingen av dem.")
        try:
            fra = tuple(int(x) for x in args.fra.split("."))
            til = tuple(int(x) for x in args.til.split("."))
            assert len(fra) == 2 and len(til) == 2
            assert 1 <= fra[1] <= 12 and 1 <= til[1] <= 12
            assert fra <= til
        except Exception:
            parser.error(f"Ugyldig periode: '{args.fra}' '{args.til}'. Forventet format: YYYY.MM")

    # Valider mappestier
    for flagg, sti in [("-F", args.faktura), ("-C", args.cloudcharge)]:
        if sti and not os.path.isdir(sti):
            parser.error(f"{flagg}: '{sti}' er ikke en gyldig mappe.")

    return fra, til, args.faktura, args.cloudcharge


if __name__ == "__main__":
    fra, til, faktura_mappe, cloudcharge_mappe = parse_args()
    beregn(fra, til, faktura_mappe, cloudcharge_mappe)
